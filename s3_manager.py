import csv
import io
import logging
import mimetypes
import os
import tempfile
from io import BytesIO, StringIO
from typing import Any, BinaryIO, Callable, Dict, List, Optional

import polars as pl
from botocore.exceptions import BotoCoreError, ClientError
from openpyxl import load_workbook

from app.core.config import settings
from .aws_manager import AWSManager

logger = logging.getLogger(__name__)
S3_BUCKET_NAME = settings.s3_bucket_name


class S3Manager(AWSManager):
    """
    S3 service manager for handling S3 operations.
    Inherits from AWSManager for common AWS configuration.
    """
    
    def __init__(self, bucket_name: Optional[str] = None, region: Optional[str] = None):
        """
        Initialize S3 manager.
        
        Args:
            bucket_name: Default bucket name for operations
            region: AWS region (defaults to AWS_REGION env var or 'us-east-1')
        """
        super().__init__('s3', region)
        self.bucket_name = bucket_name or self._get_default_bucket()
    
    def _get_default_bucket(self) -> str:
        """Get default bucket name from environment variable."""
        bucket = S3_BUCKET_NAME
        if not bucket:
            raise ValueError('S3_BUCKET_NAME must be set as an environment variable, or pass bucket_name to S3Manager.')
        return bucket
    
    def upload_file(self, file_path: str, object_key: str, bucket_name: Optional[str] = None) -> bool:
        """
        Upload a file to S3.
        
        Args:
            file_path: Local path to the file
            object_key: S3 object key (path in bucket)
            bucket_name: Bucket name (uses default if not provided)
            
        Returns:
            True if upload successful, False otherwise
        """
        bucket = bucket_name or self.bucket_name
        try:
            self.client.upload_file(file_path, bucket, object_key)
            logger.info(f"File uploaded to s3://{bucket}/{object_key}")
            return True
        except (BotoCoreError, ClientError) as e:
            logger.exception(f"Error receiving messages from SQS: {e}")
            self._handle_aws_error(e, f"upload file {file_path} to {bucket}/{object_key}")
            return False
    
    def upload_fileobj(self, file_obj: BinaryIO, object_key: str, bucket_name: Optional[str] = None) -> bool:
        """
        Upload a file object to S3.
        
        Args:
            file_obj: File-like object to upload
            object_key: S3 object key (path in bucket)
            bucket_name: Bucket name (uses default if not provided)
            
        Returns:
            True if upload successful, False otherwise
        """
        bucket = bucket_name or self.bucket_name
        try:
            content_type, _ = mimetypes.guess_type(object_key)
            extra_args = {}
            if content_type:
                extra_args["ContentType"] = content_type
            else:
                extra_args["ContentType"] = "binary/octet-stream"

            self.client.upload_fileobj(file_obj, bucket, object_key, ExtraArgs=extra_args)
            logger.info(f"File uploaded to s3://{bucket}/{object_key} (ContentType={extra_args['ContentType']})")
            return True
        except (BotoCoreError, ClientError) as e:
            self._handle_aws_error(e, f"upload file object to {bucket}/{object_key}")
            return False
    
    def multipart_upload_fileobj(self, file_obj: BinaryIO, object_key: str, bucket_name: Optional[str] = None, chunk_size: int = 5 * 1024 * 1024) -> bool:
        """
        Upload a file object to S3 using multipart upload (streaming large files).
        - chunk_size: in bytes (default 5MB, min allowed by S3)
        """
        bucket = bucket_name or self.bucket_name
        try:
            content_type, _ = mimetypes.guess_type(object_key)
            if not content_type:
                content_type = "binary/octet-stream"

            # Step 1: Create multipart upload with explicit ContentType
            response = self.client.create_multipart_upload(
                Bucket=bucket,
                Key=object_key,
                ContentType=content_type
            )
            upload_id = response['UploadId']
            logger.info(f"Multipart upload started: UploadId={upload_id}, ContentType={content_type}")

            part_number = 1
            parts = []

            while True:
                chunk = file_obj.read(chunk_size)
                if not chunk:
                    break

                part_response = self.client.upload_part(
                    Body=chunk,
                    Bucket=bucket,
                    Key=object_key,
                    UploadId=upload_id,
                    PartNumber=part_number,
                )

                parts.append({
                    'PartNumber': part_number,
                    'ETag': part_response['ETag'],
                })

                logger.debug(f"Uploaded part {part_number}, ETag={part_response['ETag']}")
                part_number += 1

            # Step 2: Complete the multipart upload
            self.client.complete_multipart_upload(
                Bucket=bucket,
                Key=object_key,
                UploadId=upload_id,
                MultipartUpload={'Parts': parts}
            )

            logger.info(f"Multipart upload completed: s3://{bucket}/{object_key}")
            return True

        except Exception as e:
            logger.error(f"Multipart upload failed: {e}")
            # Abort the multipart upload if something goes wrong
            try:
                self.client.abort_multipart_upload(
                    Bucket=bucket,
                    Key=object_key,
                    UploadId=upload_id
                )
                logger.info("Multipart upload aborted.")
            except Exception as abort_error:
                logger.error(f"Failed to abort multipart upload: {abort_error}")
            return False
    
    def download_file(self, object_key: str, file_path: str, bucket_name: Optional[str] = None) -> bool:
        """
        Download a file from S3.
        
        Args:
            object_key: S3 object key (path in bucket)
            file_path: Local path to save the file
            bucket_name: Bucket name (uses default if not provided)
            
        Returns:
            True if download successful, False otherwise
        """
        bucket = bucket_name or self.bucket_name
        try:
            self.client.download_file(bucket, object_key, file_path)
            logger.info(f"File downloaded from s3://{bucket}/{object_key} to {file_path}")
            return True
        except (BotoCoreError, ClientError) as e:
            self._handle_aws_error(e, f"download file {bucket}/{object_key} to {file_path}")
            return False
    
    def delete_object(self, object_key: str, bucket_name: Optional[str] = None) -> bool:
        """
        Delete an object from S3.
        
        Args:
            object_key: S3 object key (path in bucket)
            bucket_name: Bucket name (uses default if not provided)
            
        Returns:
            True if deletion successful, False otherwise
        """
        bucket = bucket_name or self.bucket_name
        try:
            self.client.delete_object(Bucket=bucket, Key=object_key)
            logger.info(f"Object deleted from s3://{bucket}/{object_key}")
            return True
        except (BotoCoreError, ClientError) as e:
            self._handle_aws_error(e, f"delete object {bucket}/{object_key}")
            return False
    
    def list_objects(self, prefix: str = "", bucket_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        List objects in S3 bucket with optional prefix.
        
        Args:
            prefix: Object key prefix to filter by
            bucket_name: Bucket name (uses default if not provided)
            
        Returns:
            List of object information dictionaries
        """
        bucket = bucket_name or self.bucket_name
        try:
            response = self.client.list_objects_v2(
                Bucket=bucket,
                Prefix=prefix
            )
            objects = response.get('Contents', [])
            logger.info(f"Listed {len(objects)} objects from s3://{bucket}/{prefix}")
            return objects
        except (BotoCoreError, ClientError) as e:
            self._handle_aws_error(e, f"list objects in {bucket}/{prefix}")
            return []
    
    def generate_presigned_url(self, object_key: str, expiration: int = 3600, 
                              operation: str = 'get_object', bucket_name: Optional[str] = None) -> Optional[str]:
        """
        Generate a presigned URL for S3 object access.
        
        Args:
            object_key: S3 object key (path in bucket)
            expiration: URL expiration time in seconds (default: 1 hour)
            operation: S3 operation ('get_object', 'put_object', etc.)
            bucket_name: Bucket name (uses default if not provided)
            
        Returns:
            Presigned URL or None if generation failed
        """
        bucket = bucket_name or self.bucket_name
        try:
            url = self.client.generate_presigned_url(
                operation,
                Params={'Bucket': bucket, 'Key': object_key},
                ExpiresIn=expiration
            )
            logger.info(f"Generated presigned URL for s3://{bucket}/{object_key}")
            return url
        except (BotoCoreError, ClientError) as e:
            self._handle_aws_error(e, f"generate presigned URL for {bucket}/{object_key}")
            return None
    
    def test_connection(self) -> bool:
        """
        Test S3 connection by listing buckets.
        
        Returns:
            True if connection is successful, False otherwise
        """
        try:
            self.client.list_buckets()
            return True
        except Exception as e:
            self._handle_aws_error(e, "connection test")
            return False 

    def read_tabular_file(self, file_type: str, object_key: str, bucket_name: Optional[str] = None) -> Optional[pl.DataFrame]:
        """
        Read a CSV or Excel file from S3 into a Polars DataFrame.

        Args:
            object_key: S3 object key (path in bucket)
            file_type: Type of file to read ('csv' or 'excel')
            bucket_name: Optional S3 bucket name (defaults to self.bucket_name)

        Returns:
            Polars DataFrame if successful, None otherwise.
        """
        bucket = bucket_name or self.bucket_name
        try:
            response = self.client.get_object(Bucket=bucket, Key=object_key)
            content = response['Body'].read()

            if file_type.lower() == 'csv':
                decoded = content.decode('utf-8')
                df = pl.read_csv(StringIO(decoded))
            elif file_type.lower() == 'excel':
                # Polars doesn't natively read Excel, so we use openpyxl to parse
                # and then convert rows into a Polars DataFrame.
                wb = load_workbook(BytesIO(content), read_only=True)
                ws = wb.active if wb.active is not None else wb[wb.sheetnames[0]]

                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header_row:
                    logger.error(f"No header row found in Excel file: {object_key}")
                    return None

                headers = list(header_row)
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(row)

                df = pl.DataFrame(rows, schema=headers)
            else:
                logger.error(f"Unsupported file type: {file_type}")
                return None

            logger.info(f"Tabular file read from s3://{bucket}/{object_key} using Polars")
            return df

        except (BotoCoreError, ClientError, Exception) as e:
            self._handle_aws_error(e, f"read {file_type} from {bucket}/{object_key}")
            return None
    
    async def read_csv_excel_file(
        self,
        file_type: str,
        object_key: str,
        batch_size: int = 500,
        row_callback: Optional[Callable[[List[Dict[str, Any]]], None]] = None,
        bucket_name: Optional[str] = None,
        local_file: bool = False,
    ) -> None:
        if row_callback is None:

            def default_callback(batch: List[Dict[str, Any]]) -> None:
                logger.info(f"Processed batch of {len(batch)} rows")

            row_callback = default_callback

        tmp_file_path = None

        try:
            # Decide whether to use local file or download from S3
            if local_file:
                file_path = object_key
            else:
                with tempfile.NamedTemporaryFile(
                    delete=False, suffix=os.path.splitext(object_key)[1]
                ) as tmp_file:
                    bucket = bucket_name or self.bucket_name
                    self.client.download_fileobj(bucket, object_key, tmp_file)
                    file_path = tmp_file.name


            print(f"Processing file: {file_path}")
            tmp_file_path = file_path  # store temp file path for cleanup
            # Process CSV
            if file_type.lower() == "csv":
                with open(file_path, mode="r", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    batch = []
                    for row in reader:
                        batch.append(row)
                        if len(batch) >= batch_size:
                            await row_callback(batch)
                            batch = []
                    if batch:
                        await row_callback(batch)

            # Process Excel
            elif file_type.lower() == "excel":
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active if wb.active is not None else wb[wb.sheetnames[0]]
                if ws is None:
                    raise ValueError(f"No sheet found in Excel file: {file_path}")

                header_row = next(
                    ws.iter_rows(min_row=1, max_row=1, values_only=True), None
                )
                if not header_row:
                    raise ValueError(f"No header row found in Excel file: {file_path}")

                headers = list(header_row)
                batch = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    batch.append(row_dict)
                    if len(batch) >= batch_size:
                        await row_callback(batch)
                        batch = []
                if batch:
                    await row_callback(batch)

            else:
                raise ValueError(f"Unsupported file type: {file_type}")

            logger.info(f"Finished processing {object_key}")

        finally:
            # Only delete if it’s a temp file we downloaded
            if tmp_file_path and os.path.exists(tmp_file_path):
                try:
                    os.remove(tmp_file_path)
                    logger.debug(f"Temporary file deleted: {tmp_file_path}")
                except Exception as e:
                    logger.warning(
                        f"Could not delete temporary file {tmp_file_path}: {e}"
                    )





    # def read_csv_excel_file(
    #     self,
    #     file_type: str,
    #     object_key: str,
    #     batch_size: int = 500,
    #     row_callback: Optional[Callable[[List[Dict[str, Any]]], None]] = None,
    #     bucket_name: Optional[str] = None,
    # ) -> None:
    #     """
    #     Stream a CSV or Excel file from S3 in batches and process rows incrementally.

    #     Args:
    #         file_type: 'csv' or 'excel'
    #         object_key: S3 object key (path in bucket)
    #         batch_size: Number of rows per batch
    #         row_callback: Function to call with each batch (List[Dict[str, Any]])
    #         bucket_name: Optional S3 bucket name (defaults to self.bucket_name)
    #     """
    #     bucket = bucket_name or self.bucket_name

    #     if row_callback is None:

    #         def default_callback(batch: List[Dict[str, Any]]) -> None:
    #             logger.info(f"Processed batch of {len(batch)} rows")

    #         row_callback = default_callback

    #     try:
    #         response = self.client.get_object(Bucket=bucket, Key=object_key)
    #         body = response["Body"]

    #         if file_type.lower() == "csv":
    #             # Stream CSV from S3 without loading into memory
    #             stream = io.TextIOWrapper(body, encoding="utf-8")
    #             reader = csv.DictReader(stream)

    #             batch = []
    #             for row in reader:
    #                 batch.append(row)
    #                 if len(batch) >= batch_size:
    #                     row_callback(batch)
    #                     batch = []
    #             if batch:
    #                 row_callback(batch)

    #         elif file_type.lower() == "excel":
    #             # Download Excel and use read_only mode for memory efficiency
    #             file_stream = io.BytesIO(body.read())
    #             wb = load_workbook(file_stream, read_only=True)

    #             if not wb.sheetnames:
    #                 raise ValueError(f"No sheets found in Excel file: {object_key}")

    #             ws = wb.active
    #             if ws is None:
    #                 raise ValueError(
    #                     f"Active sheet not found in Excel file: {object_key}"
    #                 )

    #             header_row = next(
    #                 ws.iter_rows(min_row=1, max_row=1, values_only=True), None
    #             )
    #             if not header_row:
    #                 raise ValueError(f"No header row found in Excel file: {object_key}")

    #             headers = list(header_row)
    #             batch = []

    #             for row in ws.iter_rows(min_row=2, values_only=True):
    #                 row_dict = dict(zip(headers, row))
    #                 batch.append(row_dict)
    #                 if len(batch) >= batch_size:
    #                     row_callback(batch)
    #                     batch = []
    #             if batch:
    #                 row_callback(batch)

    #         else:
    #             raise ValueError(f"Unsupported file type: {file_type}")

    #         logger.info(f"Finished processing s3://{bucket}/{object_key}")

    #     except Exception as e:
    #         self._handle_aws_error(e, f"stream tabular file from {bucket}/{object_key}")

